# -*- coding: utf-8 -*-
"""Monta o Excel de contatos do WhatsApp a partir do CSV bruto exportado."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "."  # pasta com o CSV exportado
ENTRADA = f"{BASE}/contatos-whatsapp-bruto.csv"
SAIDA = f"{BASE}/CONTATOS-WHATSAPP-2026-08-22.xlsx"

DDDS_VALIDOS = set(
    list(range(11, 20)) + [21, 22, 24, 27, 28] + list(range(31, 39))
    + list(range(41, 50)) + [51, 53, 54, 55] + list(range(61, 70))
    + [71, 73, 74, 75, 77, 79] + list(range(81, 90)) + list(range(91, 100))
)

DDI_PAISES = {
    "966": "Arábia Saudita", "880": "Bangladesh", "243": "RD Congo",
    "92": "Paquistão", "62": "Indonésia", "61": "Austrália",
    "44": "Reino Unido", "31": "Holanda", "27": "África do Sul",
    "1": "EUA/Canadá",
}


def classificar(tel):
    """Retorna (tipo, ddd, pais)."""
    if tel.startswith("55") and 12 <= len(tel) <= 13:
        ddd = tel[2:4]
        if int(ddd) in DDDS_VALIDOS:
            resto = tel[4:]
            if len(tel) == 13 and resto[0] == "9":
                return "Celular BR", ddd, "Brasil"
            if len(tel) == 12 and resto[0] in "6789":
                return "Celular BR (formato antigo)", ddd, "Brasil"
            if len(tel) == 12 and resto[0] in "2345":
                return "Fixo BR", ddd, "Brasil"
            return "BR (verificar)", ddd, "Brasil"
        return "Serviço/0800 BR", "", "Brasil"
    if tel.startswith("55"):
        return "BR (verificar)", "", "Brasil"
    for pref in sorted(DDI_PAISES, key=len, reverse=True):
        if tel.startswith(pref):
            return "Internacional", "", DDI_PAISES[pref]
    return "Internacional", "", "Outro"


def formatar(tel, tipo):
    if tipo.startswith("Celular BR") or tipo == "Fixo BR":
        ddd, num = tel[2:4], tel[4:]
        meio = len(num) - 4
        return f"+55 ({ddd}) {num[:meio]}-{num[meio:]}"
    return "+" + tel


# ---------- parse ----------
linhas = open(ENTRADA, encoding="utf-8").read().splitlines()
assert linhas[0].startswith("Nome;"), "cabecalho inesperado"
registros, invalidos, vistos = [], 0, set()
for ln in linhas[1:]:
    if not ln.strip():
        continue
    partes = ln.split(";")
    nome = partes[0].strip()
    m = re.search(r'="(\d+)"', ln)
    if not m:
        invalidos += 1
        continue
    tel = m.group(1)
    business = partes[-1].strip()
    if tel == "0" or len(tel) < 8:
        invalidos += 1
        continue
    if tel in vistos:
        continue
    vistos.add(tel)
    tipo, ddd, pais = classificar(tel)
    registros.append({
        "nome": nome, "tel": tel, "tipo": tipo, "ddd": ddd, "pais": pais,
        "business": "Sim" if business == "Sim" else "Não",
        "obs": "" if nome else "Sem nome",
        "fmt": formatar(tel, tipo),
    })

celulares = [r for r in registros if r["tipo"].startswith("Celular BR")]
outros = [r for r in registros if not r["tipo"].startswith("Celular BR")]

# ---------- estilos ----------
FONTE = "Arial"
LARANJA = "E67E22"
hdr_font = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=LARANJA)
cel_font = Font(name=FONTE, size=10)
borda = Border(bottom=Side(style="hair", color="DDDDDD"))


def cabecalho(ws, titulos, larguras):
    for i, (t, w) in enumerate(zip(titulos, larguras), 1):
        c = ws.cell(row=1, column=i, value=t)
        c.font, c.fill = hdr_font, hdr_fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def linha(ws, r, valores, tel_cols):
    for i, v in enumerate(valores, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font, c.border = cel_font, borda
        if i in tel_cols:
            c.number_format = "@"


wb = Workbook()

# ---------- aba Resumo ----------
res = wb.active
res.title = "Resumo"
res.column_dimensions["A"].width = 44
res.column_dimensions["B"].width = 14
t = res.cell(row=1, column=1, value="CONTATOS WHATSAPP — EXPORT 22/08/2026")
t.font = Font(name=FONTE, size=14, bold=True, color=LARANJA)
res.cell(row=2, column=1, value="Fonte: export do WhatsApp Web em 22/08/2026 (script de console)").font = Font(name=FONTE, size=9, italic=True)

itens = [
    ("Total de contatos válidos e únicos", len(registros)),
    ("Celulares Brasil (prontos para disparo)", len(celulares)),
    ("  • formato atual (9 dígitos)", sum(1 for c in celulares if c["tipo"] == "Celular BR")),
    ("  • formato antigo (8 dígitos)", sum(1 for c in celulares if "antigo" in c["tipo"])),
    ("Fixos Brasil", sum(1 for r in registros if r["tipo"] == "Fixo BR")),
    ("Serviços/0800", sum(1 for r in registros if r["tipo"] == "Serviço/0800 BR")),
    ("Internacionais", sum(1 for r in registros if r["tipo"] == "Internacional")),
    ("Contatos sem nome salvo", sum(1 for r in registros if r["obs"])),
    ("Contas WhatsApp Business", sum(1 for r in registros if r["business"] == "Sim")),
]
r = 4
for rotulo, formula in itens:
    res.cell(row=r, column=1, value=rotulo).font = cel_font
    c = res.cell(row=r, column=2, value=formula)
    c.font = Font(name=FONTE, size=10, bold=True)
    r += 1

notas = [
    "",
    "Notas:",
    f"• {invalidos} registro(s) descartado(s) na limpeza (telefone vazio ou '0').",
    "• Duplicados por telefone já removidos no export e nesta planilha.",
    "• A coluna 'Salvo na agenda' do export veio toda 'Nao' (o WhatsApp não expôs a flag neste método) e foi omitida.",
    "• A data da última conversa NÃO está neste export — será adicionada com a 2ª extração (script de conversas recentes).",
    "• Aba 'Disparo Celular BR' = somente celulares brasileiros, prontos para importar na ferramenta de disparo.",
]
for n in notas:
    res.cell(row=r, column=1, value=n).font = Font(name=FONTE, size=9)
    r += 1

# ---------- aba Todos os Contatos ----------
ws = wb.create_sheet("Todos os Contatos")
cabecalho(ws, ["Nome", "Telefone", "Tipo", "DDD", "País", "Business", "Observação", "Formatado"],
          [38, 16, 26, 7, 16, 10, 12, 22])
for i, reg in enumerate(sorted(registros, key=lambda x: (x["obs"] == "Sem nome", x["nome"].casefold())), 2):
    linha(ws, i, [reg["nome"], reg["tel"], reg["tipo"], reg["ddd"], reg["pais"],
                  reg["business"], reg["obs"], reg["fmt"]], {2})
ws.auto_filter.ref = f"A1:H{len(registros)+1}"

# ---------- aba Disparo Celular BR ----------
ws2 = wb.create_sheet("Disparo Celular BR")
cabecalho(ws2, ["Nome", "Telefone", "Formato antigo?", "Sem nome?"], [38, 16, 15, 11])
cels = sorted(celulares, key=lambda x: (x["obs"] == "Sem nome", x["nome"].casefold()))
for i, reg in enumerate(cels, 2):
    linha(ws2, i, [reg["nome"], reg["tel"],
                   "Sim" if "antigo" in reg["tipo"] else "",
                   "Sim" if reg["obs"] else ""], {2})
ws2.auto_filter.ref = f"A1:D{len(cels)+1}"

# ---------- aba Fixos-Internacional-Outros ----------
ws3 = wb.create_sheet("Fixos-Internacional-Outros")
cabecalho(ws3, ["Nome", "Telefone", "Tipo", "País"], [38, 16, 26, 16])
outs = sorted(outros, key=lambda x: (x["tipo"], x["nome"].casefold()))
for i, reg in enumerate(outs, 2):
    linha(ws3, i, [reg["nome"], reg["tel"], reg["tipo"], reg["pais"]], {2})
ws3.auto_filter.ref = f"A1:D{len(outs)+1}"

wb.save(SAIDA)

print(f"registros validos unicos: {len(registros)}")
print(f"celulares BR: {len(celulares)}  (antigos: {sum(1 for c in celulares if 'antigo' in c['tipo'])})")
print(f"outros: {len(outros)}")
print(f"sem nome: {sum(1 for r in registros if r['obs'])}")
print(f"descartados: {invalidos}")
print(f"salvo em: {SAIDA}")
